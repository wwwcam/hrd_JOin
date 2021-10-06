from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from random import randint
import random
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import os.path
import re
import pyautogui as pag
import pyperclip




def By_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
        return True
    except:
        return False
    

def By_selector(css):
    try:
        driver.find_element_by_css_selector(css)
        return True
    except:
        return False
    

def hrd_join():    
   # 초성 리스트. 00 ~ 18
   CHOSUNG_LIST = ['ㄱ', 'ㄲ', 'ㄴ', 'ㄷ', 'ㄸ', 'ㄹ', 'ㅁ', 'ㅂ', 'ㅃ', 'ㅅ', 'ㅆ', 'ㅇ', 'ㅈ', 'ㅉ', 'ㅊ', 'ㅋ', 'ㅌ', 'ㅍ', 'ㅎ']
   # 중성 리스트. 00 ~ 20
   JUNGSUNG_LIST = ['ㅏ', 'ㅐ', 'ㅑ', 'ㅒ', 'ㅓ', 'ㅔ', 'ㅕ', 'ㅖ', 'ㅗ', 'ㅘ', 'ㅙ', 'ㅚ', 'ㅛ', 'ㅜ', 'ㅝ', 'ㅞ', 'ㅟ', 'ㅠ', 'ㅡ', 'ㅢ', 'ㅣ']
   # 종성 리스트. 00 ~ 27 + 1(1개 없음)
   JONGSUNG_LIST = [' ', 'ㄱ', 'ㄲ', 'ㄳ', 'ㄴ', 'ㄵ', 'ㄶ', 'ㄷ', 'ㄹ', 'ㄺ', 'ㄻ', 'ㄼ', 'ㄽ', 'ㄾ', 'ㄿ', 'ㅀ', 'ㅁ', 'ㅂ', 'ㅄ', 'ㅅ', 'ㅆ', 'ㅇ', 'ㅈ', 'ㅊ', 'ㅋ', 'ㅌ', 'ㅍ', 'ㅎ']

def korean_to_be_englished(korean_word): 
    r_lst = []
    for w in list(korean_word.strip()): 
        ## 영어인 경우 구분해서 작성함. 
        if '가'<=w<='힣': 
            ## 588개 마다 초성이 바뀜. 
            ch1 = (ord(w) - ord('가'))//588
            ## 중성은 총 28가지 종류
            ch2 = ((ord(w) - ord('가')) - (588*ch1)) // 28
            ch3 = (ord(w) - ord('가')) - (588*ch1) - 28*ch2
            r_lst.append([CHOSUNG_LIST[ch1], JUNGSUNG_LIST[ch2], JONGSUNG_LIST[ch3]])
        else: 
            r_lst.append([w])
    return r_lst
    
def korean_word_to_initials(korean_word): 
   """
   한글을 입력받아서 한글 초성에 따라서 이니셜로 변환해줍니다.
   한국 성의 경우 조금 다르게 변환되는데 '박' ==> 'Park'인 부분은 반영하지 않음 
   """
   w_to_k = {'ㄱ': 'K', 'ㄲ': 'G', 'ㄴ': 'N', 'ㄷ': 'D', 'ㄸ': 'D', 'ㄹ': 'R', 'ㅁ': 'M', 'ㅂ': 'B', 
           'ㅃ': 'B', 'ㅅ': 'S', 'ㅈ': 'J', 'ㅉ': 'J', 'ㅊ': 'C', 'ㅌ': 'T', 'ㅍ': 'P', 'ㅎ': 'H'}
   r_lst = []
   for i, w in enumerate(korean_to_be_englished(korean_word)): 
     if w[0] in w_to_k.keys(): 
         r_lst.append( w_to_k[w[0]] )
     else: 
         if w[1] in ['ㅑ', 'ㅕ', 'ㅛ', 'ㅠ', 'ㅖ']:  
             r_lst.append('Y')
         elif w[1] in ['ㅝ', 'ㅘ', 'ㅙ', 'ㅚ', 'ㅜ', 'ㅞ', 'ㅟ']: 
             r_lst.append('W')
         elif w[1] in ['ㅔ', 'ㅡ', 'ㅢ']: 
             r_lst.append('E')
         elif w[1] in ['ㅏ', 'ㅐ']: 
             r_lst.append('A')
         elif w[1] in ['ㅓ']: 
             r_lst.append('U')
         elif w[1] in ['ㅗ']: 
             r_lst.append('O')
         elif w[1] in ['ㅣ']: 
             if i==0:  
                 r_lst.append('L')
             else: 
                 r_lst.append('I')
         else: 
             return 'not applicable'
   return "".join(r_lst)

def potp(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs): 
   #필수 (1)
   link = driver.find_element_by_css_selector('#clauses1')
   driver.execute_script("arguments[0].click();", link)
   time.sleep(0.1)       
   driver.execute_script('window.scrollTo(100,200)')
   #필수 (2)
   link = driver.find_element_by_css_selector('#clauses2')
   driver.execute_script("arguments[0].click();", link)
   driver.execute_script('window.scrollTo(200,1000)')
   #가입 (3)
   link = driver.find_element_by_css_selector('#clauses3')
   driver.execute_script("arguments[0].click();", link)
   #가입 (4)
   link = driver.find_element_by_css_selector('#clauses4')
   driver.execute_script("arguments[0].click();", link)
   #다음페이지 
   link = driver.find_element_by_css_selector('#content > div.control > button.next')
   driver.execute_script("arguments[0].click();", link)
   time.sleep(1)

def hdpselec(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs): 
   driver.switch_to.window(driver.window_handles[-1])
   print("휴대폰인증")
   time.sleep(1)
   if tongsinsa == "sk" : 
      #sk
      while By_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(1) > label') == False :
            print(tongsinsa,"기다려")
            time.sleep(0.5)
            driver.switch_to.window(driver.window_handles[-1])
      link = driver.find_element_by_css_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(1) > label')
      driver.execute_script("arguments[0].click();", link)
   elif tongsinsa == "kt" : 
      #kt        
      while By_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(2) > label > span') == False :
            print(tongsinsa,"기다려 ??? ")
            time.sleep(1)
            driver.switch_to.window(driver.window_handles[-1])
      link = driver.find_element_by_css_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(2) > label')
      driver.execute_script("arguments[0].click();", link)
   elif tongsinsa == "lg" : 
      #lg
      while By_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(3) > label') == False :
            print(tongsinsa,"기다려")
            time.sleep(0.5)
            driver.switch_to.window(driver.window_handles[-1])
      link = driver.find_element_by_css_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(3) > label')
      driver.execute_script("arguments[0].click();", link)
   elif tongsinsa == "알뜰" : 
      #알뜰
      while By_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(4) > label') == False :
            print(tongsinsa,"기다려")
            time.sleep(0.5)
            driver.switch_to.window(driver.window_handles[-1])
      link = driver.find_element_by_css_selector('#ct > form:nth-child(1) > fieldset > ul.agency_select__items > li:nth-child(4) > label')
      driver.execute_script("arguments[0].click();", link)
      time.sleep(1)
      #시작하기
      link = driver.find_element_by_css_selector('#ct > fieldset > button')
      driver.execute_script("arguments[0].click();", link)
      if tel2 == "sk" : 
         #sk
         link = driver.find_element_by_css_selector('#wrap > div.layerPopupWrap > div.layer-pop.agency_select__popup > form > div.pop-con_02 > ul > li.first-item > div.licensee_title > a > label > span.ele.sk > img')
         driver.execute_script("arguments[0].click();", link)
      elif rtel == "kt" : 
         #kt
         link = driver.find_element_by_css_selector('#wrap > div.layerPopupWrap > div.layer-pop.agency_select__popup > form > div.pop-con_02 > ul > li:nth-child(2) > div.licensee_title > a > label > span.ele.kt > img')
         driver.execute_script("arguments[0].click();", link)
      elif rtel == "lg" : 
         #lg
         link = driver.find_element_by_css_selector('#wrap > div.layerPopupWrap > div.layer-pop.agency_select__popup > form > div.pop-con_02 > ul > li.active > div.licensee_title > a > label > span.ele.lgu > img')
         driver.execute_script("arguments[0].click();", link)
      #선택
      link = driver.find_element_by_css_selector('#btnSelect')
      driver.execute_script("arguments[0].click();", link)

   #전체동의 
   link = driver.find_element_by_css_selector('#ct > form:nth-child(1) > fieldset > ul.agreelist.all > li > span > label:nth-child(2)')
   driver.execute_script("arguments[0].click();", link)
   #문자인선택
   link = driver.find_element_by_css_selector('#btnSms')
   driver.execute_script("arguments[0].click();", link)
      
def secur(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs):
   while True :
      #이름
      input_box = driver.find_element_by_css_selector('#userName')
      driver.execute_script(f"arguments[0].value = '{nm}';", input_box)
      #생년원일 1
      input_box = driver.find_element_by_css_selector('#birthDay1')
      driver.execute_script(f"arguments[0].value = '{jumin}';", input_box)
      #생년원일 2
      input_box = driver.find_element_by_css_selector('#birthDay2')
      driver.execute_script(f"arguments[0].value = '{sex}';", input_box)
      #휴대폰번호 
      input_box = driver.find_element_by_css_selector('#No')
      driver.execute_script(f"arguments[0].value = '{hdp}';", input_box)
      #보안문자  
      input_box = driver.find_element_by_css_selector('#secur')
      inp = pag.prompt('보안문자 5자리를 입력하세요:   ')
      driver.execute_script(f"arguments[0].value = \'{inp}\';", input_box)
      #확인
      link = driver.find_element_by_css_selector('#btnSubmit')
      driver.execute_script("arguments[0].click();", link)
      time.sleep(0.5)
      #재시도
      if By_selector('body > div.layer-alert-alert > div.pop-btn-alert > button') == True :
         pag.alert("보안문자가 틀리셨습니다","경고 재시도 합니다")
         pag.sleep(0.5)
         link = driver.find_element_by_css_selector('body > div.layer-alert-alert > div.pop-btn-alert > button')
         driver.execute_script("arguments[0].click();", link)
      else :
         break

def otp(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs): 
   while True : 
         #인증번호
         inp = pag.prompt('인증번호 6자리를 입력하세요:   ')
         input_box = driver.find_element_by_css_selector('#otp')
         driver.execute_script(f"arguments[0].value = \'{inp}\';", input_box)
         time.sleep(0.3)
         #확인
         link = driver.find_element_by_css_selector('#btnSubmit')
         driver.execute_script("arguments[0].click();", link)
         time.sleep(0.5)
         #재시도
         if By_selector('#btnCancel') == True :
            pag.alert("인증문자가 틀리셨습니다","경고 재시도 합니다")
            pag.sleep(0.5)
            link = driver.find_element_by_css_selector('#btnCancel')
            driver.execute_script("arguments[0].click();", link)
         else :
            break


def idcrt(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs): 
   driver.switch_to.window(driver.window_handles[-1])
   #id생성 쓰기
   input_box = driver.find_element_by_css_selector('#selIntgId')
   name = [nm]
   for case in nm: 
     print("{} ==> {}".format(case, korean_word_to_initials(case)))
   idc = (korean_word_to_initials(nm).lower()) + ('10203456789')
   driver.execute_script(f"arguments[0].value =\'{idc}\';", input_box)
   print(idc)

   #id중복확인
   link = driver.find_element_by_css_selector('#intgIdSelFrm > div > ul > li:nth-child(2) > p:nth-child(3) > button')
   driver.execute_script("arguments[0].click();", link)
   time.sleep(1)
   #id 저장 
   load_ws[f'J{i}'].value = f'{idc}'
   load_wb.save(filepath)

   #다음페이지
   link = driver.find_element_by_css_selector('#content > div.control > button.next')
   driver.execute_script("arguments[0].click();", link)
   time.sleep(1)
   #alert확인
   alert = driver.switch_to.alert.accept()
   time.sleep(2)

def one_id_input(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs): 
   #비밀번호
   input_box = driver.find_element_by_css_selector('#pwd')
   driver.execute_script(f"arguments[0].value = '{opw}';", input_box)
   driver.execute_script('window.scrollTo(200,1000)')
   input_box = driver.find_element_by_css_selector('#pwdCfm')
   driver.execute_script(f"arguments[0].value = '{opw}';", input_box)
   #핸드폰 
   input_box = driver.find_element_by_css_selector('#mobileNo1')
   hp1 = load_ws[f'G{i}'].value[0: 3]
   driver.execute_script(f"arguments[0].value = '{hp1}';", input_box)
   input_box = driver.find_element_by_css_selector('#mobileNo2')
   hp2 = load_ws[f'G{i}'].value[4: 8]
   driver.execute_script(f"arguments[0].value = '{hp2}';", input_box)
   input_box = driver.find_element_by_css_selector('#mobileNo3')
   hp3 = load_ws[f'G{i}'].value[9: 13]
   driver.execute_script(f"arguments[0].value = '{hp3}';", input_box)
   #이메일 
   input_box = driver.find_element_by_css_selector('#email')
   driver.execute_script(f"arguments[0].value = '{email}';", input_box)
   #다음페이지
   link = driver.find_element_by_css_selector('#intgUserInfoReg > div.control > button.next')
   driver.execute_script("arguments[0].click();", link)
   #alert확인
   alert = driver.switch_to.alert.accept()
   time.sleep(2)

def login(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs):
    #id
    input_box = driver.find_element_by_css_selector('#userloginId')
    driver.execute_script(f"arguments[0].value = '{id_inp}';", input_box)
    time.sleep(0.5)
    #pw
    input_box = driver.find_element_by_css_selector('#userloginPwd')
    driver.execute_script(f"arguments[0].value = '{opw}';", input_box)
    #로그인
    link = driver.find_element_by_css_selector('#loginBtn')
    driver.execute_script("arguments[0].click();", link)
    time.sleep(0.5)



if __name__ == "__main__" :
   
   sttNum = 4
   lastNum = 32

   for i in range(sttNum, lastNum): 
      driver = webdriver.Chrome(executable_path='chromedriver')
      driver.get('http://ims.work.go.kr/linkApi.do?siteId=HRDNET00&linkApiToken=JdTdApbJU7lRy2sUyd4VdNCbXwrAvDJ4tRJzIDtwIN9VUIRw4tZ4iBy8tkQBt0Y4Ji%2BSbmloG83xeUe8G1o6dA0xrt3Q1XTJ6JZYrUy3BJC4e60AAHNG%2Bl4GbP%2F73s8z1sNOGU7cMkaE5RpzMQhEi0EIAD8dI2NF2wPG6zOhkjWVZ6%2BPenFln0hugLvpGe5sC4M1cqKA%2BrCjBJn1y5s4MVrQiO3TDWwWlVTa763WLJU5n5lLDmRzxN4qkbCG2aDxNd1rM0Xp1dmmqGh%2BT1Ue8R3AGIhBfkSG9zo2GBYT1a37jKnkyWgWS4rzsEi2ZsinsK9t1rqY1UKgHTPa9UwLZbmpiLIITcPmjIc%2BxtkQBUi4oti0KtKq%2BCOe20ushHPyvomD7pJcIlabTZNAfTD2c7Cvbda6mNVCoB0z2vVMC2V6pf1bpszBIBnJU8sqJ%2BGWw5qOfuhjN94su0Jso5ifRqA2GpvyR%2Bf2khtWQE%2FrNrPlwhzW%2B%2FEiBIIIWIWHPRxj')
      #def super_get(url): 
      #    driver.get(url)
      #driver.execute_script("window.onbeforeunload = function() {};")
      #driver.execute_script("window.alert = function() {};")

      tb = time.time()
      time.sleep(0.1)

      print('HRD 가입을 시작합니다! 핸드폰을 꺼내주세요!')
      #i = input('훈련생의 번호를 입력하시요:   ')
      #tel = input('통신사를 입력해주세요:    ')

      #load_wb = load_workbook("직무교육총괄.xlsx", data_only=True)
      #load_ws = load_wb['관리대장']      
      # nm = load_ws[f'D{i}'].value ; jumin = load_ws[f'E{i}'].value[0:6] ; sex = load_ws[f'E{i}'].value ; sex = sex[7] 
      #hpNum = load_ws[f'F{i}'].value ; hdp = re.sub('-','',hpNum) ; tongsinsa = load_ws[f'K{i}'].value ; tel = load_ws[f'L{i}'].value 
      #id_inp = load_ws[f'M{i}'].value ; pw_inp = load_ws[f'N{i}'].value ; adrs = load_ws[f'O{i}'].value ; opw = "123456a!@" ; email = "wwwcam@naver.com"
      filepath = "로뎀직무교육 신청명단 양식-2021.xlsx"
      load_wb = load_workbook(filepath, data_only=True)
      load_ws = load_wb['Sheet1']

   
      nm = load_ws[f'E{i}'].value ; jumin = load_ws[f'F{i}'].value[0:6] ; sex = load_ws[f'F{i}'].value ; sex = sex[7] 
      hpNum = load_ws[f'G{i}'].value ; hdp = re.sub('-','',hpNum) ; tongsinsa = load_ws[f'H{i}'].value ; tel = load_ws[f'I{i}'].value 
      id_inp = load_ws[f'J{i}'].value ; pw_inp = load_ws[f'K{i}'].value ; adrs = load_ws[f'L{i}'].value ; opw = "123456a!@" ; email = "wwwcam@naver.com"
      '''nm = input('이름을 입력하시요:   ')
      jumin = input('생년월일을 입력하시요:   ')
      sex = input('성별을 입력하시요:   ')
      hp = input('통신사 입력하시요:   ')
      hdp = input('스마트폰 번호를 입력하시요:   ')'''

      #스크롤 후 클릭 
      link = driver.find_element_by_css_selector('#content > div.control > button.next')
      driver.execute_script("arguments[0].scrollIntoView();", link)
      driver.execute_script("arguments[0].click();", link)
      time.sleep(0.3)
      #개인정보
      potp(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
      time.sleep(0.3)
      #휴대폰인증선택 
      link = driver.find_element_by_css_selector('#content > div.article.cfm-type > ul > li.phone > div > button')
      driver.execute_script("arguments[0].click();", link)
      time.sleep(0.3)
      #핸드폰인증
      hdpselec(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
      time.sleep(0.3)
      #보안문자
      secur(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
      time.sleep(0.3)
      #인증문자
      otp(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
      time.sleep(0.3)
      
      driver.switch_to.window(driver.window_handles[-1])
      #아이디가 있습니다!
      if By_selector('#content > div.control > button') == True :
            pag.alert("아이디가 존재합니다!","경고 재시도 합니다")
            driver.switch_to.window(driver.window_handles[-1])
            odlId = driver.find_element_by_css_selector("#content > div.article > p > em")
            pag.confirm(f'이전 아이디 입니다.\n{odlId.text}', '알림')
            pag.sleep(0.5)
            #id 저장 
            load_ws[f'J{i}'].value = odlId.text
            #link = driver.find_element_by_css_selector('#content > div.control > button')
            #driver.execute_script("arguments[0].click();", link)

            input(odlId.text)
            time.sleep(0.3)
            id_inp = odlId.text.replace('"',"")
            #비밀번호 검색 
            driver.get('https://www.hrd.go.kr/hrdp/mb/pmbco/PMBCO0200D.do')
            #아이디 넣기 
            input_box = driver.find_element_by_css_selector('#loginId')
            driver.execute_script(f"arguments[0].value = \'{id_inp}\';", input_box)
            #휴대폰 본인인증 검색 
            input_box = driver.find_element_by_css_selector('#mobileCheck')
            driver.execute_script("arguments[0].click();", input_box)

            #핸드폰인증
            hdpselec(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
            #보안문자
            secur(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)    
            #인증문자
            otp(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
            time.sleep(0.5)
            driver.switch_to.window(driver.window_handles[-1])
            #새로운 pw
            input_box = driver.find_element_by_css_selector('#pass1')
            driver.execute_script(f"arguments[0].value = \'{opw}\';", input_box)
            input_box = driver.find_element_by_css_selector('#pass2')  
            driver.execute_script(f"arguments[0].value = \'{opw}\';", input_box)
            #변경하기 
            link = driver.find_element_by_css_selector('#btnModifyPass')
            driver.execute_script("arguments[0].click();", link)
            print(f"ID : {id_inp}      PW : {opw} 변경했습니다")
            time.sleep(0.5)
            alert = driver.switch_to.alert
            alert.accept()
            time.sleep(3)
            #로그인 
            link = driver.find_element_by_css_selector('#subContentWrap > div.subContArea > div > div.btnGroup.mt60.btnGroupMoTy1 > button.btnType1.primary')
            driver.execute_script("arguments[0].click();", link)
            login(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
            input("ENTER 치면 quit!!")
            driver.quit()
            print(time.time() - tb)       
            print("-------------------------------------------------------------")
            print("-------------------------------------------------------------")
      else :
         #id생성
         idcrt(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)
         #개인정보입력
         one_id_input(driver,i,nm,jumin,tongsinsa,tel, id_inp,pw_inp,adrs)

         input("다른 학생 준비해주세요!")
         print(time.time() - tb)
         driver.quit()

         input("다른 학생 준비해주세요! 준비되었으면 ENTER을 치세요!!")
         print('\n')
         print("-------------------------------------------------------------")
         print("-------------------------------------------------------------")


         print('\n')



